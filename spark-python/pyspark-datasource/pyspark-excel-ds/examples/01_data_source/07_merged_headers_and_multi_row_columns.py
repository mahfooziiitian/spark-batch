"""Merged cells & multi-row (hierarchical) column headers.

Key concepts:
    - Finance/regional templates often merge a top "category" row across
      several sub-metric columns (e.g. "Sales" merged over "Q1"/"Q2")
    - pandas.read_excel(header=[0, 1]) parses these into a two-level
      MultiIndex; Spark DataFrames need flat column names, so flatten
      before bridging with spark.createDataFrame()
    - A merged cell that spans both header rows (e.g. a single "Region"
      column) surfaces as "Unnamed: N_level_1" for its second-row label —
      detect and drop that placeholder instead of concatenating it in
"""

import openpyxl
import pandas as pd

from pys_excel import get_spark, print_dataframe, print_header, print_warning, set_log_level, temp_excel_path
from pys_excel._logging import get_logger

set_log_level("DEBUG")
logger = get_logger("example.merged_headers")


def _write_sample_with_merged_header(path: str) -> None:
    """Build a workbook with real merged header cells (as a business analyst would)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Regional"

    ws["A1"] = "Region"
    ws.merge_cells("A1:A2")
    ws["B1"] = "Sales"
    ws.merge_cells("B1:C1")
    ws["D1"] = "HR"
    ws.merge_cells("D1:E1")

    ws["B2"] = "Q1"
    ws["C2"] = "Q2"
    ws["D2"] = "Headcount"
    ws["E2"] = "Budget"

    rows = [
        ("North", 120000, 135000, 42, 5_000_000),
        ("South", 98000, 101000, 30, 3_200_000),
    ]
    for i, row in enumerate(rows, start=3):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    wb.save(path)


if __name__ == "__main__":
    spark = get_spark("excel-merged-headers")

    workbook = temp_excel_path("merged_headers_demo")
    _write_sample_with_merged_header(workbook)

    print_header("1. Read with a two-row (hierarchical) header")
    pdf = pd.read_excel(workbook, header=[0, 1], engine="openpyxl")
    logger.debug("Raw MultiIndex columns: %s", pdf.columns.tolist())

    print_header("2. Flatten the MultiIndex into Spark-friendly column names")
    flat_columns = [top if str(sub).startswith("Unnamed") else f"{top}_{sub}" for top, sub in pdf.columns]
    pdf.columns = flat_columns
    print_warning(f"Flattened columns: {flat_columns}")

    df = spark.createDataFrame(pdf)
    print_dataframe(df, title="Flattened Multi-Header Data")

    spark.stop()
